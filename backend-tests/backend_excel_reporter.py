import os
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class BackendExcelReporter:
    def __init__(self, filename="Backend_API_300_Test_Report.xlsx"):
        self.filename = filename
        self.wb = openpyxl.Workbook()
        
    def generate_report(self, test_results, start_time, end_time):
        total_tests = len(test_results)
        passed = sum(1 for r in test_results if r['status'] == 'PASS')
        failed = sum(1 for r in test_results if r['status'] == 'FAIL')
        skipped = sum(1 for r in test_results if r['status'] == 'SKIP')
        pass_rate = (passed / total_tests * 100) if total_tests > 0 else 0.0
        total_duration = end_time - start_time
        avg_latency = sum(r['duration'] for r in test_results) / total_tests if total_tests > 0 else 0.0

        # --- Sheet 1: Backend Executive Summary Dashboard ---
        ws_dash = self.wb.active
        ws_dash.title = "Backend Executive Summary"
        ws_dash.views.sheetView[0].showGridLines = True

        # Header Title Banner
        ws_dash.merge_cells("A1:G2")
        title_cell = ws_dash["A1"]
        title_cell.value = "Smart Ambulance System - Backend API Test Report (300 Test Cases)"
        title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Summary Metrics Cards
        headers_metrics = [
            ("Total API Tests", total_tests, "1E293B"),
            ("Passed Endpoints", passed, "15803D"),
            ("Failed Endpoints", failed, "B91C1C" if failed > 0 else "475569"),
            ("Skipped Tests", skipped, "D97706" if skipped > 0 else "475569"),
            ("API Pass Rate", f"{pass_rate:.2f}%", "166534"),
            ("Avg API Response Time", f"{avg_latency*1000:.2f} ms", "0284C7")
        ]

        row_start = 4
        for idx, (label, val, color) in enumerate(headers_metrics, start=1):
            col_letter = get_column_letter(idx)
            cell_lbl = ws_dash[f"{col_letter}{row_start}"]
            cell_lbl.value = label
            cell_lbl.font = Font(name="Calibri", size=10, bold=True, color="475569")
            cell_lbl.alignment = Alignment(horizontal="center", vertical="center")
            cell_lbl.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

            cell_val = ws_dash[f"{col_letter}{row_start+1}"]
            cell_val.value = val
            cell_val.font = Font(name="Calibri", size=14, bold=True, color=color)
            cell_val.alignment = Alignment(horizontal="center", vertical="center")
            cell_val.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )
            cell_lbl.border = thin_border
            cell_val.border = thin_border

        # Module Breakdown Table
        ws_dash["A7"] = "Backend Module API Pass Rate Summary"
        ws_dash["A7"].font = Font(name="Calibri", size=12, bold=True, color="0F172A")

        table_headers = ["Backend Module", "API Tests Run", "Passed", "Failed", "Skipped", "Module Pass Rate"]
        for col_idx, th in enumerate(table_headers, start=1):
            cell = ws_dash.cell(row=8, column=col_idx, value=th)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        modules = {}
        for r in test_results:
            mod = r['module']
            if mod not in modules:
                modules[mod] = {'total': 0, 'pass': 0, 'fail': 0, 'skip': 0}
            modules[mod]['total'] += 1
            if r['status'] == 'PASS':
                modules[mod]['pass'] += 1
            elif r['status'] == 'FAIL':
                modules[mod]['fail'] += 1
            else:
                modules[mod]['skip'] += 1

        curr_row = 9
        for mod_name, data in modules.items():
            mod_pass_rate = (data['pass'] / data['total'] * 100) if data['total'] > 0 else 0.0
            ws_dash.cell(row=curr_row, column=1, value=mod_name).alignment = Alignment(horizontal="left")
            ws_dash.cell(row=curr_row, column=2, value=data['total']).alignment = Alignment(horizontal="center")
            ws_dash.cell(row=curr_row, column=3, value=data['pass']).alignment = Alignment(horizontal="center")
            ws_dash.cell(row=curr_row, column=4, value=data['fail']).alignment = Alignment(horizontal="center")
            ws_dash.cell(row=curr_row, column=5, value=data['skip']).alignment = Alignment(horizontal="center")
            
            rate_cell = ws_dash.cell(row=curr_row, column=6, value=f"{mod_pass_rate:.2f}%")
            rate_cell.alignment = Alignment(horizontal="center")
            rate_cell.font = Font(bold=True, color="166534" if mod_pass_rate == 100 else "991B1B")
            
            for c in range(1, 7):
                ws_dash.cell(row=curr_row, column=c).border = Border(
                    left=Side(style='thin', color='E2E8F0'),
                    right=Side(style='thin', color='E2E8F0'),
                    top=Side(style='thin', color='E2E8F0'),
                    bottom=Side(style='thin', color='E2E8F0')
                )
            curr_row += 1

        for col in ws_dash.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_dash.column_dimensions[col_letter].width = max(max_len + 4, 16)

        # --- Sheet 2: Backend API Detailed Test Results ---
        ws_details = self.wb.create_sheet(title="Backend API Detailed Results")
        ws_details.views.sheetView[0].showGridLines = True

        detail_headers = [
            "Test ID", "Backend Module", "Test Name", "Endpoint & Method", 
            "Expected Code", "Actual Code", "Status", "Duration (s)", "Timestamp", "API Payload / Response Log"
        ]
        
        ws_details.append(detail_headers)
        for col_num in range(1, len(detail_headers) + 1):
            cell = ws_details.cell(row=1, column=col_num)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        pass_font = Font(color="15803D", bold=True)
        
        fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        fail_font = Font(color="B91C1C", bold=True)

        for row_idx, res in enumerate(test_results, start=2):
            ws_details.append([
                res['id'],
                res['module'],
                res['name'],
                res.get('endpoint', 'GET /api/health'),
                res.get('expected_code', 200),
                res.get('actual_code', 200),
                res['status'],
                round(res['duration'], 4),
                res['timestamp'],
                res['details']
            ])

            status_cell = ws_details.cell(row=row_idx, column=7)
            if res['status'] == 'PASS':
                status_cell.fill = pass_fill
                status_cell.font = pass_font
            else:
                status_cell.fill = fail_fill
                status_cell.font = fail_font

            ws_details.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            ws_details.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center")
            ws_details.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center")
            ws_details.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center")
            ws_details.cell(row=row_idx, column=8).alignment = Alignment(horizontal="right")
            ws_details.cell(row=row_idx, column=9).alignment = Alignment(horizontal="center")

            for col_idx in range(1, 11):
                ws_details.cell(row=row_idx, column=col_idx).border = Border(
                    left=Side(style='thin', color='F1F5F9'),
                    right=Side(style='thin', color='F1F5F9'),
                    top=Side(style='thin', color='F1F5F9'),
                    bottom=Side(style='thin', color='F1F5F9')
                )

        ws_details.freeze_panes = "A2"

        for col in ws_details.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_details.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 65)

        os.makedirs(os.path.dirname(os.path.abspath(self.filename)), exist_ok=True)
        self.wb.save(self.filename)
        print(f"[BackendExcelReporter] Backend API Report saved to '{self.filename}' ({total_tests} API tests).")
